# from tkinter import filedialog
# from pandas import DataFrame as pd
# import tkinter.ttk as ttk
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from os.path import isfile, join, dirname, basename, isdir  # splitext, split
from os import listdir, walk  # ,remove
from zipfile import ZipFile
from shutil import copyfileobj
from openpyxl.styles import Font
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
import re
from pandas import DataFrame
from cnpj_format import Cnpj

# TODO: Save a Path String (For portability sake)
rfid_sheet = r"local_rfid.xlsx"
# rfid_sheet = r"local_rfid.xlsx"

cols_use = [
    "Sequencia",
    "Data",
    "Emitir Por",
    "Estabelecimento",
    "Nº Nota Sensormatic (Recebida)",  # this is important
    "Pedido",
    "CNPJ",
    "Fornecedor",  # Unused on display
]
# TODO:
# Append to path when unzipping (unzipping contents) Done
# Control the append to excel...

xml_iterator = "{http://www.portalfiscal.inf.br/nfe}"


def unzip_contents(dir_path):
    def grab_zip_paths(dir_path):
        append_path = []
        # Calling filenames (List)
        _, _, filenames = next(walk(dir_path), (None, None, []))
        # Unpacking the list elements to append to basepath
        for unpack_files in filenames:
            append_path.append(
                dir_path + "\\" + unpack_files
            )  # worked better than extend
        return append_path

    finally_it = grab_zip_paths(dir_path)

    for item in finally_it:
        if item.endswith(".zip"):
            with ZipFile(item) as zip_file:
                for member in zip_file.namelist():
                    filename = basename(member)
                    # skip directories
                    if not filename:
                        continue

                    # copy file (taken from zipfile's extract)
                    # the extra truck here is shutil
                    source = zip_file.open(member)
                    target = open(join(dir_path, filename), "wb")
                    with source, target:
                        copyfileobj(source, target)
    return finally_it  # Given i want to remove the file afterwards


def search_xml_key_custom(rootget, tag):  # for fields requiring special attention
    if tag == "CNPJ":  # redirect search to [0][0][2]
        for elem in rootget[0][0][2].iter(tag=f"{xml_iterator}{tag}"):
            return elem.text
    elif tag == "infCpl":  # if i could make a first and second pass here...
        for elem in rootget[0][0].iter(tag=f"{xml_iterator}{tag}"):
            retrieve = elem.text

            retrieve_forn = retrieve[
                retrieve.rfind("ornecedor: ") : retrieve.rfind(". ")
            ].lstrip("ornecedor: ")

            retrieve_logr = retrieve[
                retrieve.rfind("Endereco Entrega:") : retrieve.rfind("| ")
            ].lstrip("Endereco Entrega: ")
            retrieve_f = f"{retrieve_forn} - {retrieve_logr}"
            return retrieve_f

    elif tag == "vIPI":
        for elem in rootget[0][0][4].iter(tag=f"{xml_iterator}{tag}"):
            return elem.text
    else:
        for elem in rootget[0][0].iter(tag=f"{xml_iterator}{tag}"):
            return elem.text


def search_xml_key(rootget, tag):
    # This starts mid-rootget (on the other function)
    for elem in rootget[0][0].iter(tag=f"{xml_iterator}{tag}"):
        return elem.text


def format_date(dt_str):
    dt_str_t = f"{dt_str[8:10]}/{dt_str[5:7]}/{dt_str[0:4]}"
    return dt_str_t


def establ(lookupstring):
    associated_fornecs = {
        "45242914004518": "CDT 020",
        "45242914038250": "CDSC 093",
        "45242914004275": "CDR 050",
    }

    for word, initial in associated_fornecs.items():
        lookupstring = lookupstring.replace(word, initial)

    return lookupstring


# This might break if logr differentiates?
""" def filiate_sort(stringlookup):
    # by data on geographic addresses
    beira_rio_fornecs = {
        "R DARCI AZAMBUJA, 555, CENTRO, ROCA SALES - RS ": "(005)",
        "RUA CRUZEIRO DO SUL, 856 PAV 02, VL NOVA, SAPIRANGA - RS": "(006)",
        "RUA JOAO PAULO I, 77, CENTRO, IGREJINHA - RS": "(003)",
        "ST VER OSCAR HORN, 1520, CANUDOS, NOVO HAMBURGO - RS": "(???)",
    }

    if stringlookup.startswith("CALCADOS BEIRA RIO"):
        for word, initial in beira_rio_fornecs.items():
            stringlookup = stringlookup.replace(word, initial)
    else:  # cleaning if not
        sub_ind_sl = stringlookup.find("- ")
        stringlookup = stringlookup[:sub_ind_sl]

    return stringlookup """


def read_xml_into_obj(xml_object, mask_enable="False"):
    data_volume = []
    with open(xml_object) as xml_object:
        tree = ET.parse(xml_object)
        rootget = tree.getroot()

        emitter = search_xml_key_custom(rootget, "CNPJ")
        nnf = search_xml_key(rootget, "nNF")
        dh_emi = search_xml_key(rootget, "dhEmi")
        inf_cpl = search_xml_key_custom(rootget, "infCpl")
        valor_unit = search_xml_key(rootget, "vUnCom")
        item = search_xml_key(rootget, "xProd")
        valor_ipi = search_xml_key(rootget, "vIPI")
        pedido_ca = search_xml_key(rootget, "infCpl")
        qtd = search_xml_key(rootget, "qCom")
        total_nota = search_xml_key(rootget, "vNF")

        # Cleaning and Filtering
        # inf_cpl = inf_cpl.replace(".", "")  # To remove annoying dots after all
        # inf_cpl = filiate_sort(inf_cpl)
        # inf_cpl = inf_cpl.split(",")  # to make sure cnpj doesn't appear
        # inf_cpl = inf_cpl[0]
        name_forn = inf_cpl.split(",")[0]
        # Find the sequence of 14 digits anywhere on string
        if mask_enable == "True":
            cnpj_seek = re.search(r"\d{14}", inf_cpl).group(0)
            cnpj_forn = Cnpj().format(cnpj_seek)  # this is slowing down my program
        else:
            cnpj_forn = re.search(r"\d{14}", inf_cpl).group(0)

        # 05891905000112. Endereco Entrega

        emitter = establ(emitter)  # ok?
        valor_ipi = valor_ipi.replace(".", ",")
        dh_emi = dh_emi.split("T")
        dh_emi = dh_emi[0]
        dh_emi = format_date(dh_emi)

        pedido_ca = pedido_ca[-9:]  # Not subscriptable

        # extra cleaning the ipi_field, for 1.333.00 cases
        valor_unit = valor_unit[::-1].replace(".", ",", 1)
        valor_unit = valor_unit[::-1].replace(".", "")
        total_nota = total_nota.replace(".", ",")
        qtd = qtd.replace(".", ",")

        data_volume.extend(
            [
                nnf,  # 0
                item,  # 1
                qtd,  # 2
                valor_unit,  # 3
                total_nota,  # 4
                name_forn,  # 5
                valor_ipi,  # 6
                pedido_ca,  # 7
                dh_emi,  # 8
                emitter,  # 9
                cnpj_forn,  # 10
            ]
        )
        return data_volume


# For some reason, it doesn't reach to populate the fields...
""" def populate_fields(
    frame_in, num_nota, item, qtd, val_unit, total, forn, ipi, ped, ind_val
):
    # Use index from combobox instead (And have it as the same event on other widgets
    # ind_val  # start at here
    frame_in.index += 1
    END = "end"
    for index, row in frame_in.iterrows():
        if index == ind_val:  # Next and previous might be incr/decr for the pandas base
            num_nota.insert(END, row["Nota Entrada"])  # same as cols_use
            item.insert(END, row["Item"])
            qtd.insert(END, row["Quantidade"])
            val_unit.insert(END, row["Val_Unitario"])
            total.insert(END, row["Total Nota"])
            forn.insert(END, row["Fornecedor"])
            ipi.insert(END, row["IPI"])
            ped.insert(END, row["Pedido"]) """


""" def initialize_combo_boxes(frame_in, num_nota, forn, ped, ind_val):
    initialized = False
    if num_nota
    ind_val = num_nota.currentIndex()  # start at here
    frame_in.index += 1
    base_list_nnf = (
        frame_in["Nota Entrada"].astype(str).tolist()
    )  # Convert to string list
    base_list_nnf.insert(0, "<Select...>")
    num_nota.addItems(base_list_nnf)
    base_list_forn = frame_in["Fornecedor"].astype(str).tolist()
    base_list_forn.insert(0, "<Select...>")  # Because lists always start with zero
    forn.addItems(base_list_forn)
    base_list_ped = frame_in["Pedido"].astype(str).tolist()
    base_list_ped.insert(0, "<Select...>")
    ped.addItems(base_list_ped)
    initialized = True
    return initialized """


# Better approach in progress
def populate_fields(
    frame_in, num_nota, item, qtd, val_unit, cnpj_item, total, forn, ipi, ped, ind_val
):
    # Use index from combobox instead (And have it as the same event on other widgets

    frame_in.index += 1
    # num_nota["values"] = frame_in[["Nota Entrada"]].values.tolist()
    # num_nota.addItems(frame_in["Nota Entrada"].astype(str).tolist())
    # forn.addItems(frame_in[["Fornecedor"]].values.tolist())
    # ped.addItems(frame_in[["Pedido"]].values.tolist())

    num_nota.blockSignals(True)
    forn.blockSignals(True)
    ped.blockSignals(True)

    num_nota.clear()
    forn.clear()
    ped.clear()

    base_list_nnf = (
        frame_in["Nota Entrada"].astype(str).tolist()
    )  # Convert to string list
    base_list_nnf.insert(0, "<Select...>")
    num_nota.addItems(base_list_nnf)
    base_list_forn = frame_in["Fornecedor"].astype(str).tolist()
    base_list_forn.insert(0, "<Select...>")  # Because lists always start with zero
    forn.addItems(base_list_forn)
    base_list_ped = frame_in["Pedido"].astype(str).tolist()
    base_list_ped.insert(0, "<Select...>")
    ped.addItems(base_list_ped)

    for index, row in frame_in.iterrows():
        if index == ind_val:  # Next and previous might be incr/decr for the pandas base
            # if index == num_nota.current():
            num_nota.setCurrentIndex(index)
            print("Index inside populate", index)
            item.setText(row["Item"])
            qtd.setText(row["Quantidade"])
            val_unit.setText(row["Val_Unitario"])
            cnpj_item.setText(row["CNPJ"])
            total.setText(row["Total Nota"])
            forn.setCurrentIndex(index)
            ipi.setText(row["IPI"])
            ped.setCurrentIndex(index)

    num_nota.blockSignals(False)
    forn.blockSignals(False)
    ped.blockSignals(False)


""" def unzip_and_pack(pick_it):  # Maybe i can reuse the first... But this is for the tk
    return unzip_contents(pick_it) """


def read_in_xml_data(pick_folder):
    # Generates an output to Pandas and also Excel
    pick_xml = []
    append_data = []

    # unzip_contents(pick_folder)  # No need to zip when reading
    # for path, dir_list, file_list in walk
    if pick_folder:
        if isfile(pick_folder):
            path = dirname(pick_folder)  # Unused in this block
            base = basename(pick_folder)[:-4]  # Takes away name
            pick_xml.append(base + ".xml")  # To assemble here? wut?

        elif isdir(pick_folder):
            path = pick_folder
            base = basename(pick_folder)

            for file in listdir(pick_folder):
                if file[-4:] == ".xml":
                    pick_xml.append(file)

    for index in range(len(pick_xml)):
        xml_file = pick_xml[index]
        xml_object = join(path + "/" + xml_file)  # Such a criptic mess
        # Gonna have to clean variables and better understand this
        reading = read_xml_into_obj(xml_object, "True")
        # reading = clean_data_volume(reading)
        append_data.append(reading)

    return append_data


def exception_handler(condition):
    if condition:
        raise Exception("Workbook Not Found")


def save_xml_to_excel(excel_data: list):
    # TODO: Maybe i should put the Exception Handler here?
    #
    def applying_styles(data):
        for c in data:
            # if c == 1:
            c = Cell(ws, column="A", row=1, value=c)
            c.font = Font(name="Arial", size=12)
            c.alignment = Alignment(horizontal="center")
            yield c

    def get_contents(data_append):  # yield some dataframe here
        # filtered_grid = DataFrame(data=data_append, cols=cols_use)
        return_append = []
        for each_row in data_append:
            empty_field = [""]
            data_append = [
                empty_field[0],
                each_row[8],  # dh_emi
                empty_field[0],
                each_row[9],  # emitter
                int(each_row[0]),  # nnf
                each_row[7],  # Pedido
                (each_row[10]),  # cnpj
                each_row[5],  # infcpl(forn)
            ]
            return_append += [data_append]
        return return_append

    wb = load_workbook(rfid_sheet)
    # exception_handler(wb)
    ws = wb["Data_Feed"]
    tab = ws.tables["Table1"]
    coord = list(range_boundaries(tab.ref))

    extracted_list = get_contents(excel_data)
    sorted_frame = DataFrame(data=extracted_list, columns=cols_use)
    sorted_frame = sorted_frame.sort_values(
        by=["Fornecedor"],
        ascending=True,
        inplace=False,
        ignore_index=True,
    )

    for row in sorted_frame.values:  # After all, playing the row game
        coord[-1] += 1
        ws.append(applying_styles(row))
        tab.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"
    wb.save(rfid_sheet)
    """ except FileNotFoundError as e:
        print("Workbook not Found!{}".format(e)) """
