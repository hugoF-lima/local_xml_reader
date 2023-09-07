import openpyxl

# import datetime
import automate_synchro_collect
import automate_synchro_main
from itertools import count
from openpyxl.styles import Font, Alignment

# import numbers

# from openpyxl.styles.numbers import NumberFormatDescriptor

from openpyxl.styles import Font, Alignment

""" from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell """

from read_xml_data import rfid_sheet as source_data

# source_data = r"src\02-Feb_2023\synchro_automation_rev_2\local_rfid.xlsx"


def return_recent_row(excel_file):
    # Compatible only with first enumerated sequences
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)

    # Select the first worksheet
    ws = wb["Data_Feed"]

    # Identify the table
    table = ws.tables["Table1"]

    # Get the data from the "Seq" column
    seq_data = [row[0].value for row in ws[table.ref][1:]]

    # Find the highest value in the "Seq" column
    highest_seq = max(seq_data)

    # Filter the table to only show rows with the highest value in the "Seq" column
    filtered_table = [row for row in ws[table.ref][1:] if row[0].value == highest_seq]
    return wb, filtered_table


# Maybe it would be safer to write a single row per instance
# extracts what will be handled by Pandas in the fillout.
def extract_data_from_table(filter_data):
    append_data = []
    for row in filter_data:
        append_data.append([cell.value for cell in row[0:]])
    return append_data


def update_rows(excel_file, fill_match):
    def applying_styles(cell):
        cell.font = Font(name="Arial", size=12)
        cell.alignment = Alignment(horizontal="center")

    wb, filtered_table = return_recent_row(excel_file)
    for match in fill_match:
        for key, value in match.items():
            for row in filtered_table:
                if row[5].value == str(key):
                    for i, cell in enumerate(row[8:]):
                        if i < len(value):  # This checks to avoid out of index error
                            cell.value = value[i]
                            applying_styles(cell)
                            if i == 2:
                                # cell.number_format = datetime.strftime(value[i])
                                """cell.number_format = datetime.strftime(
                                    value[i].strptime("%d/%m/%Y")
                                )"""
                                cell.number_format = "dd/mm/yyyy"
                            elif i == 3:
                                # cell.number_format = "0000"
                                # cell.data_type = "n"
                                cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC

                # cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
    wb.save(source_data)
    wb.close()


""" # I wonder if this wb passing works
def update_rows(excel_file, fill_match):
    wb, filtered_table = return_recent_row(excel_file)
    for match in fill_match:
        for key, value in match.items():
            for row in filtered_table:
                if row[5].value == str(key):
                    for i, cell in enumerate(row[8:]):
                        if i < len(value):
                            cell.value = value[i]
                # cell.data_type = openpyxl.cell.cell.TYPE_NUMERIC
    wb.save(source_data)
    wb.close() """


""" def update_rows(excel_file, fill_match):
    wb, filtered_table = return_recent_row(excel_file)

    def applying_styles(cell):
        cell.font = Font(name="Arial", size=12)
        cell.alignment = Alignment(horizontal="center")

    for match in fill_match:
        for key, value in match.items():
            for row in filtered_table:
                if row[5].value == str(key):
                    for i, cell in enumerate(row[8:]):
                        cell.value = value[i]
                        # cell_value = ""

                        # Check the corresponding datetime cell
                        if i == 2:  # and isinstance(value[i], datetime.datetime):
                            # Assign descriptor (Format cell as Short Date)
                            cell.number_format = NumberFormatDescriptor("dd/mm/yyyy")
                            # cell_value = datetime.datetime.strptime(str(value[i]), "%d/%m/%Y")
                            # cell.number_format = numbers.FORMAT_DATE_XLSX15

                        # cell.value = cell_value
                        # Apply styles to the cell
                        applying_styles(cell)
    wb.save(excel_file)
    wb.close() """


# wb, filtered_tab = return_recent_row(source_data)


# cfop_val = "6.949"
# fill_match = automate_synchro.collect_data_from_note(cfop_val)
# new_data = extract_data_from_table(filtered_tab)
# update_rows(source_data, fill_match)


def collect_data(w_get, notes_ahead, data_emit, cfop_val, widget_spin):
    # validate if excel file is available first
    automate_synchro_main.window_restore(w_spotted=w_get, module_str="FISCAL")

    if notes_ahead > 0:
        for number in count(start=1):
            fill_match = automate_synchro_collect.collect_data_from_note(
                cfop_val, data_emit
            )
            widget_spin.stepDown()
            update_rows(source_data, fill_match)
            if number == notes_ahead:
                break


""" wb, filtered_table = return_recent_row(source_data)
# for match in fill_match:
# for key, value in match.items():
for row in filtered_table:
    print(row[5].value) """


# append_filtered_data = []
# Print the filtered table

# Gather data from table first:

# append_filtered_data.append([cell.value for cell in row[0:]])

# wb.save(source_data)
